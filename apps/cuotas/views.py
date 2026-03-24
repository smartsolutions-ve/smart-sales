"""Vistas del módulo de cuotas y ventas."""
import csv
from decimal import Decimal, InvalidOperation
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.views.decorators.http import require_http_methods, require_POST
from django.db.models import Sum, F
from django.http import HttpResponse

from apps.accounts.decorators import role_required
from .models import VentaMensual, TasaCambio, Zona


@login_required
@role_required('gerente')
def lista(request):
    qs = VentaMensual.objects.filter(organization=request.org)

    # Filtros
    periodo = request.GET.get('periodo', '')
    zona = request.GET.get('zona', '')
    vendedor = request.GET.get('vendedor', '')
    canal = request.GET.get('canal', '')

    if periodo:
        qs = qs.filter(periodo=periodo + '-01')
    if zona:
        qs = qs.filter(zona_nombre__icontains=zona)
    if vendedor:
        qs = qs.filter(vendedor_nombre__icontains=vendedor)
    if canal:
        qs = qs.filter(canal=canal)

    zonas = VentaMensual.objects.filter(organization=request.org).values_list('zona_nombre', flat=True).distinct().order_by('zona_nombre')
    vendedores = VentaMensual.objects.filter(organization=request.org).values_list('vendedor_nombre', flat=True).distinct().order_by('vendedor_nombre')

    return render(request, 'cuotas/lista.html', {
        'ventas': qs,
        'periodo_filtro': periodo,
        'zona_filtro': zona,
        'vendedor_filtro': vendedor,
        'canal_filtro': canal,
        'zonas': zonas,
        'vendedores': vendedores,
    })


@login_required
@role_required('gerente')
def resumen_zona(request):
    qs = VentaMensual.objects.filter(organization=request.org)

    periodo = request.GET.get('periodo', '')
    if periodo:
        qs = qs.filter(periodo=periodo + '-01')

    datos = (
        qs.values('zona_nombre')
        .annotate(
            plan_total=Sum('plan_venta_usd'),
            real_total=Sum('real_venta_usd'),
            plan_cant=Sum('plan_cantidad'),
            real_cant=Sum('real_cantidad'),
        )
        .order_by('zona_nombre')
    )

    for d in datos:
        d['cumplimiento'] = round(float(d['real_total'] or 0) / float(d['plan_total']) * 100, 1) if d['plan_total'] else 0

    return render(request, 'cuotas/resumen_zona.html', {
        'datos': datos,
        'periodo_filtro': periodo,
    })


@login_required
@role_required('gerente')
def resumen_vendedor(request):
    qs = VentaMensual.objects.filter(organization=request.org)

    periodo = request.GET.get('periodo', '')
    if periodo:
        qs = qs.filter(periodo=periodo + '-01')

    datos = (
        qs.values('vendedor_nombre')
        .annotate(
            plan_total=Sum('plan_venta_usd'),
            real_total=Sum('real_venta_usd'),
            plan_cant=Sum('plan_cantidad'),
            real_cant=Sum('real_cantidad'),
        )
        .order_by('-real_total')
    )

    for d in datos:
        d['cumplimiento'] = round(float(d['real_total'] or 0) / float(d['plan_total']) * 100, 1) if d['plan_total'] else 0

    return render(request, 'cuotas/resumen_vendedor.html', {
        'datos': datos,
        'periodo_filtro': periodo,
    })


@login_required
@role_required('gerente')
def resumen_producto(request):
    qs = VentaMensual.objects.filter(organization=request.org)

    periodo = request.GET.get('periodo', '')
    if periodo:
        qs = qs.filter(periodo=periodo + '-01')

    datos = (
        qs.values('producto_nombre', 'codigo_producto')
        .annotate(
            plan_total=Sum('plan_venta_usd'),
            real_total=Sum('real_venta_usd'),
            plan_cant=Sum('plan_cantidad'),
            real_cant=Sum('real_cantidad'),
        )
        .order_by('-real_total')
    )

    for d in datos:
        d['cumplimiento'] = round(float(d['real_total'] or 0) / float(d['plan_total']) * 100, 1) if d['plan_total'] else 0

    return render(request, 'cuotas/resumen_producto.html', {
        'datos': datos,
        'periodo_filtro': periodo,
    })


@login_required
@role_required('gerente')
@require_http_methods(['GET', 'POST'])
def importar(request):
    resultado = None
    if request.method == 'POST':
        archivo = request.FILES.get('archivo')
        if not archivo:
            messages.error(request, 'Seleccione un archivo Excel.')
            return render(request, 'cuotas/importar.html')

        try:
            resultado = _procesar_excel(request, archivo)
            messages.success(request, f'Importación completada: {resultado["creados"]} creados, {resultado["actualizados"]} actualizados.')
        except Exception as e:
            messages.error(request, f'Error al importar: {e}')

    return render(request, 'cuotas/importar.html', {'resultado': resultado})


@login_required
@role_required('gerente')
def exportar_csv(request):
    qs = VentaMensual.objects.filter(organization=request.org)

    periodo = request.GET.get('periodo', '')
    if periodo:
        qs = qs.filter(periodo=periodo + '-01')

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="cuotas_ventas.csv"'
    response.write('\ufeff')  # BOM for Excel

    writer = csv.writer(response)
    writer.writerow([
        'Periodo', 'Vendedor', 'Producto', 'Código', 'Zona', 'Canal',
        'Plan Cant.', 'Plan Venta USD', 'Real Cant.', 'Real Venta USD',
        '% Cumplimiento',
    ])

    for v in qs:
        writer.writerow([
            v.periodo.strftime('%Y-%m'),
            v.vendedor_nombre,
            v.producto_nombre,
            v.codigo_producto,
            v.zona_nombre,
            v.canal,
            v.plan_cantidad,
            v.plan_venta_usd,
            v.real_cantidad,
            v.real_venta_usd,
            v.cumplimiento_venta,
        ])

    return response


# ── Tasas de cambio ───────────────────────────────────────────────

@login_required
@role_required('gerente')
def tasas_lista(request):
    tasas = TasaCambio.objects.filter(organization=request.org)
    return render(request, 'cuotas/tasas_lista.html', {'tasas': tasas})


@login_required
@role_required('gerente')
@require_http_methods(['GET', 'POST'])
def tasa_crear(request):
    if request.method == 'POST':
        fecha = request.POST.get('fecha', '')
        tasa = request.POST.get('tasa_bs_por_usd', '')
        fuente = request.POST.get('fuente', '').strip()

        if not fecha or not tasa:
            messages.error(request, 'Fecha y tasa son requeridos.')
            return render(request, 'cuotas/tasa_form.html')

        TasaCambio.objects.update_or_create(
            organization=request.org,
            fecha=fecha,
            fuente=fuente,
            defaults={'tasa_bs_por_usd': Decimal(tasa)},
        )
        messages.success(request, 'Tasa de cambio guardada.')
        return redirect('cuotas:tasas')

    return render(request, 'cuotas/tasa_form.html')


@login_required
@role_required('gerente')
@require_http_methods(['GET', 'POST'])
def tasa_editar(request, pk):
    tasa = get_object_or_404(TasaCambio, pk=pk, organization=request.org)
    if request.method == 'POST':
        fecha = request.POST.get('fecha', '')
        tasa_val = request.POST.get('tasa_bs_por_usd', '')
        fuente = request.POST.get('fuente', '').strip()

        if not fecha or not tasa_val:
            messages.error(request, 'Fecha y tasa son requeridos.')
            return render(request, 'cuotas/tasa_form.html', {'tasa': tasa})

        tasa.fecha = fecha
        tasa.tasa_bs_por_usd = Decimal(tasa_val)
        tasa.fuente = fuente
        tasa.save()
        messages.success(request, 'Tasa de cambio actualizada.')
        return redirect('cuotas:tasas')

    return render(request, 'cuotas/tasa_form.html', {'tasa': tasa})


@login_required
@role_required('gerente')
@require_POST
def tasa_eliminar(request, pk):
    tasa = get_object_or_404(TasaCambio, pk=pk, organization=request.org)
    tasa.delete()
    messages.success(request, 'Tasa de cambio eliminada.')
    return redirect('cuotas:tasas')


# ── Helpers ───────────────────────────────────────────────────────

def _procesar_excel(request, archivo):
    import openpyxl
    from django.db import transaction

    wb = openpyxl.load_workbook(archivo, read_only=True, data_only=True)

    # Intentar hoja "Cuotas" primero, luego la primera
    if 'Cuotas' in wb.sheetnames:
        ws = wb['Cuotas']
    else:
        ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError('El archivo está vacío.')

    # Primera fila como encabezados
    headers = [str(h).strip().lower() if h else '' for h in rows[0]]

    creados = 0
    actualizados = 0
    errores = []

    # Mapeo de columnas flexible
    col_map = _build_column_map(headers)

    with transaction.atomic():
        for i, row in enumerate(rows[1:], start=2):
            try:
                data = {key: row[idx] if idx is not None and idx < len(row) else None
                        for key, idx in col_map.items()}

                vendedor_nombre = str(data.get('vendedor') or '').strip()
                producto_nombre = str(data.get('producto') or '').strip()
                if not vendedor_nombre and not producto_nombre:
                    continue  # Fila vacía

                periodo_raw = data.get('periodo')
                if not periodo_raw:
                    continue

                # Parsear periodo
                from datetime import date, datetime
                if isinstance(periodo_raw, datetime):
                    periodo = periodo_raw.date().replace(day=1)
                elif isinstance(periodo_raw, date):
                    periodo = periodo_raw.replace(day=1)
                else:
                    # Intentar parsear YYYY-MM o YYYY-MM-DD
                    s = str(periodo_raw).strip()
                    if len(s) == 7:
                        periodo = date(int(s[:4]), int(s[5:7]), 1)
                    else:
                        periodo = datetime.strptime(s[:10], '%Y-%m-%d').date().replace(day=1)

                zona_nombre = str(data.get('zona') or '').strip()
                codigo = str(data.get('codigo') or '').strip()
                canal = str(data.get('canal') or '').strip().upper()

                def to_dec(val):
                    if val is None:
                        return Decimal('0')
                    try:
                        return Decimal(str(val))
                    except (InvalidOperation, ValueError):
                        return Decimal('0')

                # Auto-create zona
                zona_obj = None
                if zona_nombre:
                    zona_obj, _ = Zona.objects.get_or_create(
                        organization=request.org,
                        nombre=zona_nombre,
                        defaults={'codigo': ''},
                    )

                obj, created = VentaMensual.objects.update_or_create(
                    organization=request.org,
                    periodo=periodo,
                    vendedor_nombre=vendedor_nombre,
                    producto_nombre=producto_nombre,
                    zona_nombre=zona_nombre,
                    defaults={
                        'codigo_producto': codigo,
                        'canal': canal if canal in ('DIRECTO', 'DISTRIBUCIÓN') else '',
                        'distribucion': str(data.get('distribucion') or '').strip(),
                        'zona': zona_obj,
                        'plan_cantidad': to_dec(data.get('plan_cantidad')),
                        'plan_precio_usd': to_dec(data.get('plan_precio_usd')),
                        'plan_venta_usd': to_dec(data.get('plan_venta_usd')),
                        'plan_costo_usd': to_dec(data.get('plan_costo_usd')),
                        'plan_margen_usd': to_dec(data.get('plan_margen_usd')),
                        'plan_flete_usd': to_dec(data.get('plan_flete_usd')),
                        'plan_gastos_fin_usd': to_dec(data.get('plan_gastos_fin_usd')),
                        'plan_impuestos_usd': to_dec(data.get('plan_impuestos_usd')),
                        'plan_logisticos_usd': to_dec(data.get('plan_logisticos_usd')),
                        'plan_gastos_ventas_usd': to_dec(data.get('plan_gastos_ventas_usd')),
                        'plan_margen_neto_usd': to_dec(data.get('plan_margen_neto_usd')),
                        'real_cantidad': to_dec(data.get('real_cantidad')),
                        'real_venta_ves': to_dec(data.get('real_venta_ves')),
                        'real_venta_usd': to_dec(data.get('real_venta_usd')),
                        'real_costo_usd': to_dec(data.get('real_costo_usd')),
                        'real_flete_usd': to_dec(data.get('real_flete_usd')),
                        'real_gastos_fin_usd': to_dec(data.get('real_gastos_fin_usd')),
                        'real_impuestos_usd': to_dec(data.get('real_impuestos_usd')),
                        'real_logisticos_usd': to_dec(data.get('real_logisticos_usd')),
                        'real_gastos_ventas_usd': to_dec(data.get('real_gastos_ventas_usd')),
                        'real_margen_neto_usd': to_dec(data.get('real_margen_neto_usd')),
                    },
                )
                if created:
                    creados += 1
                else:
                    actualizados += 1

            except Exception as e:
                errores.append(f'Fila {i}: {e}')

    return {
        'creados': creados,
        'actualizados': actualizados,
        'errores': errores,
        'total': creados + actualizados,
    }


def _build_column_map(headers):
    """Mapea nombres de columna a índices. Acepta variaciones comunes."""
    mappings = {
        'periodo': ['periodo', 'mes', 'month', 'fecha'],
        'vendedor': ['vendedor', 'seller', 'nombre vendedor'],
        'producto': ['producto', 'product', 'nombre producto'],
        'codigo': ['codigo', 'código', 'code', 'sku', 'codigo producto'],
        'zona': ['zona', 'zone', 'territory'],
        'canal': ['canal', 'channel'],
        'distribucion': ['distribucion', 'distribución', 'distribution'],
        'plan_cantidad': ['plan cantidad', 'plan cant', 'plan_cantidad', 'cuota cantidad'],
        'plan_precio_usd': ['plan precio', 'plan precio usd', 'plan_precio_usd'],
        'plan_venta_usd': ['plan venta', 'plan venta usd', 'plan_venta_usd', 'cuota venta'],
        'plan_costo_usd': ['plan costo', 'plan costo usd', 'plan_costo_usd'],
        'plan_margen_usd': ['plan margen', 'plan margen usd', 'plan_margen_usd'],
        'plan_flete_usd': ['plan flete', 'plan flete usd', 'plan_flete_usd'],
        'plan_gastos_fin_usd': ['plan gastos fin', 'plan gastos financieros', 'plan_gastos_fin_usd'],
        'plan_impuestos_usd': ['plan impuestos', 'plan_impuestos_usd'],
        'plan_logisticos_usd': ['plan logisticos', 'plan logísticos', 'plan_logisticos_usd'],
        'plan_gastos_ventas_usd': ['plan gastos ventas', 'plan_gastos_ventas_usd'],
        'plan_margen_neto_usd': ['plan margen neto', 'plan_margen_neto_usd'],
        'real_cantidad': ['real cantidad', 'real cant', 'real_cantidad', 'venta cantidad'],
        'real_venta_ves': ['real venta ves', 'real venta bs', 'real_venta_ves', 'venta ves'],
        'real_venta_usd': ['real venta', 'real venta usd', 'real_venta_usd', 'venta usd'],
        'real_costo_usd': ['real costo', 'real costo usd', 'real_costo_usd'],
        'real_flete_usd': ['real flete', 'real_flete_usd'],
        'real_gastos_fin_usd': ['real gastos fin', 'real gastos financieros', 'real_gastos_fin_usd'],
        'real_impuestos_usd': ['real impuestos', 'real_impuestos_usd'],
        'real_logisticos_usd': ['real logisticos', 'real logísticos', 'real_logisticos_usd'],
        'real_gastos_ventas_usd': ['real gastos ventas', 'real_gastos_ventas_usd'],
        'real_margen_neto_usd': ['real margen neto', 'real_margen_neto_usd'],
    }

    col_map = {}
    for key, variants in mappings.items():
        col_map[key] = None
        for v in variants:
            if v in headers:
                col_map[key] = headers.index(v)
                break

    return col_map
